from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
SOURCE_PATH = ROOT / "fixtures" / "IT_Exec_Reporting_Ingestion_Template_v2_dummy_data.xlsx"
OUTPUT_PATH = ROOT / "fixtures" / "IT_Exec_Reporting_Ingestion_Template_v3_dummy_data.xlsx"
PUBLIC_OUTPUT_PATH = ROOT / "public" / "templates" / "IT_Exec_Reporting_Ingestion_Template_v3_dummy_data.xlsx"
TEMPLATE_KEY = "IT_EXEC_TEMPLATE_V3"
TEMPLATE_VERSION = 3
MONTHS = ["2026-01", "2026-02", "2026-03", "2026-04", "2026-05", "2026-06"]

DOMAIN_MAP = {
    "infrastructure": "Infrastructure",
    "euc": "End-user computing",
    "security": "Security & compliance",
    "data": "Applications & data",
    "product": "Product / development",
    "transformation": "Business transformation",
}

GANTT_ITEMS = [
    {
        "domain": "infrastructure",
        "name": "WAN Resilience Uplift",
        "sponsor": "Head of IT",
        "start_weeks": -4,
        "duration_weeks": 10,
        "complete_weeks": 5,
        "status_rag": "Amber",
        "detail": "Secondary path design agreed; implementation underway",
        "milestones": [("Secondary path live", 6)],
    },
    {
        "domain": "infrastructure",
        "name": "Private Cloud Rationalisation",
        "sponsor": "Head of IT",
        "start_weeks": 8,
        "duration_weeks": 12,
        "complete_weeks": 0,
        "status_rag": "Amber",
        "detail": "Commercial options review in progress",
        "milestones": [("Options decision", 12)],
    },
    {
        "domain": "euc",
        "name": "Laptop Refresh Phase 3",
        "sponsor": "COO",
        "start_weeks": -2,
        "duration_weeks": 8,
        "complete_weeks": 3,
        "status_rag": "Green",
        "detail": "Q3 refresh ready to launch; supplier slots confirmed",
        "milestones": [("Deployment start", 2)],
    },
    {
        "domain": "euc",
        "name": "Mobile Fleet Review",
        "sponsor": "IT Operations",
        "start_weeks": 6,
        "duration_weeks": 6,
        "complete_weeks": 0,
        "status_rag": "Green",
        "detail": "Scheduled for Q3; estate in good health",
        "milestones": [],
    },
    {
        "domain": "security",
        "name": "Security Remediation Sprint",
        "sponsor": "CEO",
        "start_weeks": -6,
        "duration_weeks": 8,
        "complete_weeks": 7,
        "status_rag": "Green",
        "detail": "Critical backlog cleared; wrapping up",
        "milestones": [("Closure & handover", 2)],
    },
    {
        "domain": "security",
        "name": "Access Control Rationalisation",
        "sponsor": "Security Lead",
        "start_weeks": 10,
        "duration_weeks": 10,
        "complete_weeks": 0,
        "status_rag": "Amber",
        "detail": "Cross-functional dependency; scheduling in progress",
        "milestones": [],
    },
    {
        "domain": "security",
        "name": "Patch Compliance Uplift",
        "sponsor": "Infrastructure Manager",
        "start_weeks": -8,
        "duration_weeks": 12,
        "complete_weeks": 9,
        "status_rag": "Green",
        "detail": "Continuous programme; trajectory positive",
        "milestones": [],
    },
    {
        "domain": "data",
        "name": "Data & Reporting Uplift",
        "sponsor": "CFO",
        "start_weeks": -10,
        "duration_weeks": 14,
        "complete_weeks": 11,
        "status_rag": "Green",
        "detail": "Management reporting foundation established",
        "milestones": [("Visual design pack", 3)],
    },
    {
        "domain": "data",
        "name": "MI Self-Service Reporting",
        "sponsor": "Data Lead",
        "start_weeks": 12,
        "duration_weeks": 10,
        "complete_weeks": 0,
        "status_rag": "Green",
        "detail": "Dependent on ingestion model; planned Q4",
        "milestones": [],
    },
    {
        "domain": "product",
        "name": "TABS Prioritisation Reset",
        "sponsor": "IT Director",
        "start_weeks": 2,
        "duration_weeks": 8,
        "complete_weeks": 0,
        "status_rag": "Amber",
        "detail": "Business prioritisation required before mobilisation",
        "milestones": [("Scope agreed", 4)],
    },
    {
        "domain": "product",
        "name": "Dev Backlog Quality Reset",
        "sponsor": "IT Director",
        "start_weeks": 14,
        "duration_weeks": 8,
        "complete_weeks": 0,
        "status_rag": "Amber",
        "detail": "Archive and close criteria decision pending",
        "milestones": [],
    },
    {
        "domain": "transformation",
        "name": "Portfolio Governance Refresh",
        "sponsor": "CIO",
        "start_weeks": -2,
        "duration_weeks": 6,
        "complete_weeks": 2,
        "status_rag": "Green",
        "detail": "Prioritisation discipline; steering cadence defined",
        "milestones": [("Steering cadence", 4)],
    },
    {
        "domain": "transformation",
        "name": "Identity Modernisation",
        "sponsor": "CEO",
        "start_weeks": -16,
        "duration_weeks": 16,
        "complete_weeks": 15,
        "status_rag": "Green",
        "detail": "Security uplift and admin overhead reduction delivered",
        "milestones": [("Project closure", 1)],
    },
]


def apply_table_style(table: Table) -> None:
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )


def ensure_metadata(readme_sheet) -> None:
    readme_sheet["A10"] = "Template Key"
    readme_sheet["B10"] = TEMPLATE_KEY
    readme_sheet["A11"] = "Template Version"
    readme_sheet["B11"] = TEMPLATE_VERSION
    readme_sheet["A12"] = "Network metric handling"
    readme_sheet["B12"] = "Overall Network KPI is derived from office rows in INPUT_Office_Network_Avail."
    readme_sheet["A13"] = "Portfolio Gantt handling"
    readme_sheet["B13"] = "12-week rolling Gantt is derived from INPUT_Gantt_Workstreams and INPUT_Gantt_Milestones."


def reset_sheet(workbook, title: str, index: int):
    if title in workbook.sheetnames:
        existing = workbook[title]
        workbook.remove(existing)
    return workbook.create_sheet(title, index=index)


def first_monday_on_or_after(month: str) -> date:
    year, month_number = map(int, month.split("-"))
    month_start = date(year, month_number, 1)
    weekday = month_start.weekday()
    days_to_monday = 0 if weekday == 0 else 7 - weekday
    return month_start + timedelta(days=days_to_monday)


def format_date(value: date) -> str:
    return value.isoformat()


def progress_date(start_date: date, complete_weeks: int) -> str:
    if complete_weeks <= 0:
        return ""
    return format_date(start_date + timedelta(days=(complete_weeks * 7) - 1))


def end_date(start_date: date, duration_weeks: int) -> str:
    return format_date(start_date + timedelta(days=(duration_weeks * 7) - 1))


def update_periods_sheet(workbook) -> None:
    sheet = workbook["Periods"]
    sheet["A1"] = "Reporting Periods"
    sheet["A2"] = "One row per reporting month. Mark exactly one row as the current period and provide the stored report cut-off date."
    sheet["F3"] = "Report Cut-Off Date"

    for row_index in range(4, sheet.max_row + 1):
        month_value = sheet.cell(row=row_index, column=1).value
        if not month_value:
            continue
        year, month_number = map(int, str(month_value).split("-"))
        sheet.cell(row=row_index, column=6).value = format_date(date(year, month_number, 19))


def build_gantt_workstreams_sheet(workbook) -> None:
    target_index = workbook.sheetnames.index("INPUT_Rolling_Roadmap") + 1
    sheet = reset_sheet(workbook, "INPUT_Gantt_Workstreams", target_index)

    sheet["A1"] = "INPUT Gantt Workstreams"
    sheet["A2"] = "One row per workstream per reporting month. Dates are rendered into the 12-week rolling Gantt."
    headers = [
        "Reporting Month",
        "Workstream Name",
        "Sponsor / Owner",
        "Domain",
        "Status RAG",
        "Start Date",
        "End Date",
        "Progress Date",
        "Detail / Commentary",
        "Display Order",
        "In Scope",
    ]
    sheet.append(headers)

    for month in MONTHS:
        base_date = first_monday_on_or_after(month)
        for display_order, item in enumerate(GANTT_ITEMS, start=1):
            start = base_date + timedelta(days=item["start_weeks"] * 7)
            sheet.append(
                [
                    month,
                    item["name"],
                    item["sponsor"],
                    DOMAIN_MAP[item["domain"]],
                    item["status_rag"],
                    format_date(start),
                    end_date(start, item["duration_weeks"]),
                    progress_date(start, item["complete_weeks"]),
                    item["detail"],
                    display_order,
                    "Yes",
                ]
            )

    last_row = 3 + len(MONTHS) * len(GANTT_ITEMS)
    table = Table(displayName="TPortfolioGanttWorkstreams", ref=f"A3:K{last_row}")
    apply_table_style(table)
    sheet.add_table(table)


def build_gantt_milestones_sheet(workbook) -> None:
    target_index = workbook.sheetnames.index("INPUT_Gantt_Workstreams") + 1
    sheet = reset_sheet(workbook, "INPUT_Gantt_Milestones", target_index)

    sheet["A1"] = "INPUT Gantt Milestones"
    sheet["A2"] = "Child milestone rows for the Portfolio Gantt. Multiple milestones per workstream are supported."
    headers = ["Reporting Month", "Workstream Name", "Milestone Label", "Milestone Date", "Display Order"]
    sheet.append(headers)

    for month in MONTHS:
        base_date = first_monday_on_or_after(month)
        display_order = 1
        for item in GANTT_ITEMS:
            start_date_value = base_date + timedelta(days=item["start_weeks"] * 7)
            for milestone_label, week_offset in item["milestones"]:
                sheet.append(
                    [
                        month,
                        item["name"],
                        milestone_label,
                        format_date(start_date_value + timedelta(days=week_offset * 7)),
                        display_order,
                    ]
                )
                display_order += 1

    last_row = sheet.max_row
    table = Table(displayName="TPortfolioGanttMilestones", ref=f"A3:E{last_row}")
    apply_table_style(table)
    sheet.add_table(table)


def main() -> None:
    if not SOURCE_PATH.exists():
        raise SystemExit(f"Source workbook not found: {SOURCE_PATH}")

    workbook = load_workbook(SOURCE_PATH)
    ensure_metadata(workbook["README"])
    update_periods_sheet(workbook)
    build_gantt_workstreams_sheet(workbook)
    build_gantt_milestones_sheet(workbook)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    PUBLIC_OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)
    workbook.save(PUBLIC_OUTPUT_PATH)

    print(f"Created {OUTPUT_PATH}")
    print(f"Created {PUBLIC_OUTPUT_PATH}")


if __name__ == "__main__":
    main()
