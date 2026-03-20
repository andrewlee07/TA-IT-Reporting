from __future__ import annotations

from calendar import monthrange
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
SOURCE_PATH = ROOT / "fixtures" / "IT_Exec_Reporting_Ingestion_Template_v1_dummy_data.xlsx"
OUTPUT_PATH = ROOT / "fixtures" / "IT_Exec_Reporting_Ingestion_Template_v2_dummy_data.xlsx"
PUBLIC_OUTPUT_PATH = ROOT / "public" / "templates" / "IT_Exec_Reporting_Ingestion_Template_v2_dummy_data.xlsx"
TEMPLATE_KEY = "IT_EXEC_TEMPLATE_V2"
TEMPLATE_VERSION = 2

OFFICES = [
    ("Birmingham", "Midlands", 1, 238, 255, [99.92, 99.95, 100.00, 100.00, 99.98, 100.00]),
    ("Manchester", "North West", 2, 196, 195, [99.88, 99.93, 99.97, 100.00, 100.00, 100.00]),
    ("Leeds", "Yorkshire", 3, 228, 185, [99.75, 99.80, 99.90, 99.95, 99.98, 100.00]),
    ("London City", "London", 4, 298, 355, [99.95, 99.97, 100.00, 100.00, 100.00, 100.00]),
    ("London East", "London", 5, 308, 360, [99.90, 99.93, 99.96, 99.98, 100.00, 100.00]),
    ("Bristol", "South West", 6, 148, 375, [99.85, 99.90, 99.93, 99.97, 100.00, 100.00]),
    ("Sheffield", "Yorkshire", 7, 228, 198, [99.70, 99.78, 99.88, 99.92, 99.96, 99.98]),
    ("Nottingham", "East Midlands", 8, 242, 232, [99.80, 99.85, 99.92, 99.96, 99.99, 100.00]),
    ("Leicester", "East Midlands", 9, 244, 248, [99.88, 99.91, 99.95, 99.98, 100.00, 100.00]),
    ("Cardiff", "Wales", 10, 128, 378, [97.80, 98.50, 99.20, 99.42, 99.55, 99.62]),
    ("Swansea", "Wales", 11, 102, 370, [98.10, 98.60, 99.10, 99.40, 99.58, 99.70]),
    ("Liverpool", "North West", 12, 182, 208, [99.82, 99.88, 99.93, 99.97, 100.00, 100.00]),
    ("Newcastle", "North East", 13, 228, 108, [99.78, 99.84, 99.91, 99.95, 99.98, 100.00]),
    ("Derby", "East Midlands", 14, 238, 238, [99.83, 99.89, 99.94, 99.97, 99.99, 100.00]),
    ("Coventry", "Midlands", 15, 248, 262, [99.87, 99.91, 99.96, 99.98, 100.00, 100.00]),
    ("Worcester", "Midlands", 16, 218, 272, [99.91, 99.94, 99.97, 99.99, 100.00, 100.00]),
    ("Oxford", "South East", 17, 248, 308, [99.93, 99.96, 99.98, 100.00, 100.00, 100.00]),
    ("Cambridge", "East", 18, 292, 295, [99.89, 99.93, 99.96, 99.99, 100.00, 100.00]),
    ("Southampton", "South", 19, 228, 398, [99.85, 99.89, 99.94, 99.97, 100.00, 100.00]),
    ("Plymouth", "South West", 20, 105, 455, [99.76, 99.82, 99.89, 99.94, 99.97, 100.00]),
    ("Norwich", "East", 21, 328, 295, [99.80, 99.86, 99.92, 99.96, 99.98, 100.00]),
    ("Exeter", "South West", 22, 128, 448, [99.81, 99.87, 99.92, 99.96, 99.99, 100.00]),
]

MONTHS = ["2026-01", "2026-02", "2026-03", "2026-04", "2026-05", "2026-06"]


def apply_table_style(table: Table) -> None:
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )


def ensure_metadata(readme_sheet) -> None:
    readme_sheet["A10"] = "Template Key"
    readme_sheet["B10"] = TEMPLATE_KEY
    readme_sheet["A11"] = "Template Version"
    readme_sheet["B11"] = TEMPLATE_VERSION
    readme_sheet["A12"] = "Network metric handling"
    readme_sheet["B12"] = "Overall Network KPI is derived from office rows in INPUT_Office_Network_Availability."


def promote_headers_to_row_three(sheet, title: str, subtitle: str) -> None:
    if (sheet["A1"].value or "") == title and (sheet["A2"].value or "") == subtitle:
        return

    sheet.insert_rows(1, amount=2)
    sheet["A1"] = title
    sheet["A2"] = subtitle


def reset_sheet(workbook, title: str, index: int):
    if title in workbook.sheetnames:
        existing = workbook[title]
        workbook.remove(existing)

    return workbook.create_sheet(title, index=index)


def build_office_locations_sheet(workbook) -> None:
    entities_index = workbook.sheetnames.index("Entities")
    sheet = reset_sheet(workbook, "Office_Locations", entities_index + 1)

    sheet["A1"] = "Office Locations"
    sheet["A2"] = "In-scope UK offices used for the network map and office-availability view."
    headers = ["Office Name", "Region", "In Scope", "Display Order", "Map X", "Map Y"]
    sheet.append(headers)

    for office_name, region, display_order, map_x, map_y, _ in OFFICES:
        sheet.append([office_name, region, "Yes", display_order, map_x, map_y])

    table = Table(displayName="TOfficeLocations", ref=f"A3:F{3 + len(OFFICES)}")
    apply_table_style(table)
    sheet.add_table(table)


def estimated_outage_minutes(month: str, availability_pct: float) -> int:
    year, month_number = month.split("-")
    minutes_in_month = monthrange(int(year), int(month_number))[1] * 24 * 60
    downtime_fraction = max(0, 100 - availability_pct) / 100
    return round(minutes_in_month * downtime_fraction)


def estimated_incidents(availability_pct: float) -> int:
    if availability_pct < 99:
        return 1
    return 0


def build_office_network_sheet(workbook) -> None:
    target_index = workbook.sheetnames.index("Office_Locations") + 1
    sheet = reset_sheet(workbook, "INPUT_Office_Network_Avail", target_index)

    sheet["A1"] = "INPUT Office Network Availability"
    sheet["A2"] = "One row per office per reporting month. Overall Network KPI is derived from this sheet."
    headers = ["Reporting Month", "Office Name", "Availability %", "Outage Minutes", "Major Incidents", "Commentary"]
    sheet.append(headers)

    for month_index, month in enumerate(MONTHS):
        for office_name, _, _, _, _, availabilities in OFFICES:
            availability_pct = availabilities[month_index]
            outage_minutes = estimated_outage_minutes(month, availability_pct)
            incidents = estimated_incidents(availability_pct)
            if availability_pct == 100:
                commentary = "No incidents."
            elif availability_pct >= 99.9:
                commentary = "Minor transient disruption."
            elif availability_pct >= 99:
                commentary = "Monitored connectivity issue."
            else:
                commentary = "User-impacting outage recorded."

            sheet.append(
                [
                    month,
                    office_name,
                    f"{availability_pct:.2f}%",
                    outage_minutes,
                    incidents,
                    commentary,
                ]
            )

    last_row = 3 + len(MONTHS) * len(OFFICES)
    table = Table(displayName="TOfficeNetworkAvailability", ref=f"A3:F{last_row}")
    apply_table_style(table)
    sheet.add_table(table)


def remove_manual_network_rows(workbook) -> None:
    sheet = workbook["INPUT_Service_Availability"]

    for row_index in range(sheet.max_row, 3, -1):
        if sheet.cell(row=row_index, column=2).value == "Network":
            sheet.delete_rows(row_index, 1)

    table = sheet.tables.get("TServiceAvailability")
    if table:
        table.ref = f"A3:J{sheet.max_row}"


def main() -> None:
    if not SOURCE_PATH.exists():
        raise SystemExit(f"Source workbook not found: {SOURCE_PATH}")

    workbook = load_workbook(SOURCE_PATH)
    ensure_metadata(workbook["README"])
    promote_headers_to_row_three(
        workbook["Periods"],
        "Reporting Periods",
        "One row per reporting month. Mark exactly one row as the current period.",
    )
    promote_headers_to_row_three(
        workbook["Entities"],
        "Entities",
        "Reference list for services, platforms and domains in scope for the exec pack.",
    )
    remove_manual_network_rows(workbook)
    build_office_locations_sheet(workbook)
    build_office_network_sheet(workbook)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    PUBLIC_OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)
    workbook.save(PUBLIC_OUTPUT_PATH)

    print(f"Created {OUTPUT_PATH}")
    print(f"Created {PUBLIC_OUTPUT_PATH}")


if __name__ == "__main__":
    main()
