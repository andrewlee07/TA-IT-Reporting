from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
SOURCE_PATH = ROOT / "fixtures" / "IT_Exec_Reporting_Ingestion_Template_v3_dummy_data.xlsx"
OUTPUT_PATH = ROOT / "fixtures" / "IT_Exec_Reporting_Ingestion_Template_v4_dummy_data.xlsx"
PUBLIC_OUTPUT_PATH = ROOT / "public" / "templates" / "IT_Exec_Reporting_Ingestion_Template_v4_dummy_data.xlsx"
TEMPLATE_KEY = "IT_EXEC_TEMPLATE_V4"
TEMPLATE_VERSION = 4
CHART_SETTINGS_SHEET = "INPUT_Chart_Settings"


def apply_table_style(table: Table) -> None:
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )


def reset_sheet(workbook, title: str, index: int):
    if title in workbook.sheetnames:
        workbook.remove(workbook[title])
    return workbook.create_sheet(title, index=index)


def ensure_metadata(readme_sheet) -> None:
    metadata_rows = {
        "Template Key": TEMPLATE_KEY,
        "Template Version": TEMPLATE_VERSION,
        "Chart settings handling": "Chart overlays are controlled in INPUT_Chart_Settings.",
    }

    for row_index in range(1, readme_sheet.max_row + 5):
        key = readme_sheet.cell(row=row_index, column=1).value
        if key in metadata_rows:
            readme_sheet.cell(row=row_index, column=2).value = metadata_rows.pop(key)

    next_row = readme_sheet.max_row + 1
    for key, value in metadata_rows.items():
        readme_sheet.cell(row=next_row, column=1).value = key
        readme_sheet.cell(row=next_row, column=2).value = value
        next_row += 1


def build_chart_settings_sheet(workbook) -> None:
    target_index = workbook.sheetnames.index("INPUT_Narrative_Notes") + 1
    sheet = reset_sheet(workbook, CHART_SETTINGS_SHEET, target_index)

    sheet["A1"] = "INPUT Chart Settings"
    sheet["A2"] = "Optional chart overlays and thresholds. One row per chart per reporting month."
    headers = [
        "Reporting Month",
        "Page",
        "Chart Key",
        "Overlay Enabled",
        "Overlay Metric",
        "Rolling Window",
        "Healthy Min",
        "Amber Min",
        "Commentary",
    ]
    sheet.append(headers)

    periods_sheet = workbook["Periods"]
    month_values = []
    for row_index in range(4, periods_sheet.max_row + 1):
        month_value = periods_sheet.cell(row=row_index, column=1).value
        if month_value:
            month_values.append(str(month_value))

    for month in month_values:
        sheet.append(
            [
                month,
                "Support Operations",
                "support_ticket_volumes",
                "Yes",
                "Close Balance %",
                3,
                100,
                97,
                "Overlay highlights whether ticket closures are keeping pace with incoming demand.",
            ]
        )

    table = Table(displayName="TChartSettings", ref=f"A3:I{sheet.max_row}")
    apply_table_style(table)
    sheet.add_table(table)


def main() -> None:
    if not SOURCE_PATH.exists():
        raise SystemExit(f"Source workbook not found: {SOURCE_PATH}")

    workbook = load_workbook(SOURCE_PATH)
    ensure_metadata(workbook["README"])
    build_chart_settings_sheet(workbook)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    PUBLIC_OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)
    workbook.save(PUBLIC_OUTPUT_PATH)

    print(f"Created {OUTPUT_PATH}")
    print(f"Created {PUBLIC_OUTPUT_PATH}")


if __name__ == "__main__":
    main()
